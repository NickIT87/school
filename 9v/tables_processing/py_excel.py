from openpyxl import Workbook

# Create a new workbook
wb = Workbook()

# Select the active worksheet (default is the first one)
ws = wb.active

# Write data to cells
ws['A1'] = 'Hello'
ws['B1'] = 'World'

# Save the workbook
wb.save('example.xlsx')
